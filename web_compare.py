"""
SOAP vs REST Karşılaştırma Web Arayüzü
Kullanım: python web_compare.py
Tarayıcı: http://localhost:8080
"""

import http.server
import socketserver
import json
import urllib.parse
import os
import sys
from datetime import datetime
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PORT = 8080
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXPORTS_DIR = os.path.join(SCRIPT_DIR, "exports")

# ========================== EXCEL STYLES ==========================
HEADER_FONT = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
DIFF_FONT = Font(name='Calibri', bold=True, size=11, color='FF0000')
SAME_FONT = Font(name='Calibri', size=11, color='008000')
NORMAL_FONT = Font(name='Calibri', size=11)
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='2F5496')
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=12, color='2F5496')
DIFF_BOLD_FONT = Font(name='Calibri', bold=True, size=11)
SUMMARY_HEADER_FILL = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
SUMMARY_VALUE_FONT = Font(name='Calibri', bold=True, size=11, color='C00000')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical='center')


# ========================== EXCEL HELPERS ==========================
def add_header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER


def add_data_row(ws, row, data, is_diff):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=str(val) if val is not None else '')
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGN
        if col == 4:
            cell.font = DIFF_FONT if is_diff else SAME_FONT
        else:
            cell.font = DIFF_BOLD_FONT if is_diff else NORMAL_FONT


def add_title(ws, row, title, col_count=4):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)


def add_subtitle(ws, row, title, col_count=4):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)


def set_col_widths(ws, widths=None):
    if widths is None:
        widths = {'A': 40, 'B': 45, 'C': 45, 'D': 15}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ========================== PARSERS ==========================
def parse_soap_xml(xml_content):
    """SOAP XML içeriğini parse eder."""
    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError as e:
        return {'error': f'XML Parse Hatası: {str(e)}'}

    for elem in root.iter():
        if '}' in elem.tag:
            elem.tag = elem.tag.split('}', 1)[1]

    result = {}
    gsd = None
    for elem in root.iter('GenerateSummaryDocument'):
        gsd = elem
        break
    if gsd is None:
        gsd = root

    result['operationResult'] = {}
    op_result = gsd.find('.//operationResult')
    if op_result is not None:
        rc = op_result.find('resultCode')
        rd = op_result.find('resultDescription')
        result['operationResult']['resultCode'] = rc.text if rc is not None else ''
        result['operationResult']['resultDescription'] = rd.text if rd is not None else ''

    for field in ['identifier', 'catalogProcessId', 'operationTypeId']:
        elem = gsd.find('.//' + field)
        result[field] = elem.text if elem is not None else ''

    customer_elem = gsd.find('.//customer')
    result['customer'] = {}
    if customer_elem is not None:
        for child in customer_elem:
            result['customer'][child.tag] = child.text if child.text else ''

    result['summaryDocumentUrls'] = []
    for url_elem in gsd.iter('summaryDocumentUrl'):
        if url_elem.text:
            result['summaryDocumentUrls'].append(url_elem.text)

    result['documents'] = []
    for doc_info in gsd.iter('summaryDocumentInfo'):
        doc = {}
        for fn in ['documentType', 'willBeScanned', 'printable', 'signatureRequired',
                   'minPage', 'maxPage', 'mandatory', 'sendEmail', 'barcodeCode',
                   'fromCampaign', 'requiredDocumentCode', 'requiredDocumentName', 'canBeSentLater']:
            elem = doc_info.find(fn)
            doc[fn] = elem.text if elem is not None and elem.text else ''

        template_info = doc_info.find('.//documentToSignDocumentTemplateInfo')
        if template_info is not None:
            tid = template_info.find('documentTemplateId')
            doc['templateId'] = tid.text if tid is not None else ''
            doc['templateFields'] = {}
            for field in template_info.findall('documentTemplateFields'):
                name_elem = field.find('name')
                value_elem = field.find('value')
                if name_elem is not None and name_elem.text:
                    doc['templateFields'][name_elem.text] = value_elem.text if value_elem is not None and value_elem.text else ''
        else:
            doc['templateId'] = ''
            doc['templateFields'] = {}

        doc['docStoreSystems'] = []
        for dss in doc_info.iter('docStoreSystem'):
            store = {'indexes': {}}
            ne = dss.find('name')
            store['name'] = ne.text if ne is not None else ''
            dc = dss.find('documentClass')
            store['documentClass'] = dc.text if dc is not None else ''
            for idx in dss.findall('documentIndexes'):
                in_ = idx.find('name')
                iv = idx.find('value')
                if in_ is not None and in_.text:
                    store['indexes'][in_.text] = iv.text if iv is not None and iv.text else ''
            doc['docStoreSystems'].append(store)

        result['documents'].append(doc)

    return result


def parse_rest_json(json_content):
    """REST JSON içeriğini parse eder."""
    try:
        raw = json.loads(json_content)
    except json.JSONDecodeError as e:
        return {'error': f'JSON Parse Hatası: {str(e)}'}

    content = raw.get('content', raw)
    result = {}

    op_result = content.get('operationResult', {})
    result['operationResult'] = {
        'resultCode': str(op_result.get('resultCode', '')),
        'resultDescription': op_result.get('resultDescription', '')
    }
    result['identifier'] = str(content.get('identifier', ''))
    result['catalogProcessId'] = str(content.get('catalogProcessId', ''))
    result['operationTypeId'] = str(content.get('operationTypeId', ''))

    cust = content.get('customer', {})
    result['customer'] = {}
    if cust:
        for key in ['type', 'customerNumber', 'nationality', 'citizenNumber', 'taxNumber',
                    'firstName', 'lastName', 'mobilePhone', 'birthDate', 'fullName',
                    'email', 'applicantType', 'registeredCityDistrict', 'address']:
            val = cust.get(key)
            result['customer'][key] = str(val) if val is not None else ''

    result['summaryDocumentUrls'] = content.get('summaryUrlList', []) or []

    result['documents'] = []
    for doc_raw in content.get('summaryDocumentList', []):
        doc = {}
        doc['documentType'] = str(doc_raw.get('documentType', ''))
        doc['willBeScanned'] = str(doc_raw.get('willBeScanned', '')).lower()
        doc['printable'] = str(doc_raw.get('printable', '')).lower()
        doc['signatureRequired'] = str(doc_raw.get('signatureRequired', '')).lower()
        doc['minPage'] = str(doc_raw.get('minPage', ''))
        doc['maxPage'] = str(doc_raw.get('maxPage', ''))
        doc['mandatory'] = str(doc_raw.get('mandatory', '')).lower()
        doc['sendEmail'] = str(doc_raw.get('sendEmail', '')).lower()
        doc['barcodeCode'] = str(doc_raw.get('barcodeCode', '')) if doc_raw.get('barcodeCode') else ''
        doc['fromCampaign'] = str(doc_raw.get('fromCampaign', '')).lower()
        doc['requiredDocumentCode'] = doc_raw.get('requiredDocumentCode', '')
        doc['requiredDocumentName'] = doc_raw.get('requiredDocumentName', '')
        doc['canBeSentLater'] = str(doc_raw.get('canBeSendLater', '')).lower()
        doc['templateId'] = str(doc_raw.get('templateId', ''))

        doc['templateFields'] = {}
        for field in doc_raw.get('documentTemplateFields', []):
            fname = field.get('name', '')
            fval = field.get('value')
            doc['templateFields'][fname] = str(fval) if fval is not None else ''

        doc['docStoreSystems'] = []
        for dss_raw in doc_raw.get('docStoreSystemList', []):
            store = {
                'name': dss_raw.get('name', ''),
                'documentClass': dss_raw.get('documentClass', ''),
                'indexes': {}
            }
            for idx in dss_raw.get('documentIndexes', []):
                store['indexes'][idx.get('name', '')] = str(idx.get('value')) if idx.get('value') is not None else ''
            doc['docStoreSystems'].append(store)

        result['documents'].append(doc)

    return result


# ========================== COMPARATOR ==========================
def normalize(val):
    if val is None:
        return ''
    val = str(val).strip()
    if val.lower() in ('null', 'none', ''):
        return ''
    return val


def display_val(val, source='soap'):
    """Görüntüleme için değeri formatla."""
    if val is None:
        return '*(boş)*' if source == 'soap' else 'null'
    v = str(val).strip()
    if v == '':
        return '*(boş)*'
    return v


def compare_fields(soap_val, rest_val):
    """İki değeri karşılaştır, (soap_display, rest_display, is_different) döner."""
    s_display = display_val(soap_val, 'soap')
    r_display = display_val(rest_val, 'rest')
    s_norm = normalize(soap_val)
    r_norm = normalize(rest_val)
    is_diff = s_norm != r_norm
    return s_display, r_display, is_diff


def compare_data(soap_data, rest_data):
    if 'error' in soap_data:
        return {'error': soap_data['error']}
    if 'error' in rest_data:
        return {'error': rest_data['error']}

    results = {
        'topLevel': [],
        'customer': [],
        'documents': [],
        'summary': {'total': 0, 'same': 0, 'diff': 0}
    }

    def add_row(section, field, soap_val, rest_val):
        s = normalize(soap_val)
        r = normalize(rest_val)
        is_same = s == r
        results[section].append({
            'field': field,
            'soap': soap_val if soap_val else '*(boş)*',
            'rest': rest_val if rest_val else '*(boş)*',
            'status': 'same' if is_same else 'diff'
        })
        results['summary']['total'] += 1
        if is_same:
            results['summary']['same'] += 1
        else:
            results['summary']['diff'] += 1

    # Top-Level
    for f in ['resultCode', 'resultDescription']:
        add_row('topLevel', f, soap_data.get('operationResult', {}).get(f, ''),
                rest_data.get('operationResult', {}).get(f, ''))
    for f in ['identifier', 'catalogProcessId', 'operationTypeId']:
        add_row('topLevel', f, soap_data.get(f, ''), rest_data.get(f, ''))

    # Customer
    cust_map = [
        ('type', 'type'), ('crmCustomerId', 'customerNumber'), ('nationality', 'nationality'),
        ('citizenNumber', 'citizenNumber'), ('taxNumber', 'taxNumber'), ('firstName', 'firstName'),
        ('lastName', 'lastName'), ('mobilePhone', 'mobilePhone'), ('birthDate', 'birthDate')
    ]
    soap_cust = soap_data.get('customer', {})
    rest_cust = rest_data.get('customer', {})
    for soap_key, rest_key in cust_map:
        add_row('customer', soap_key, soap_cust.get(soap_key, ''), rest_cust.get(rest_key, ''))

    # Documents
    soap_docs = {d['requiredDocumentCode']: d for d in soap_data.get('documents', [])}
    rest_docs = {d['requiredDocumentCode']: d for d in rest_data.get('documents', [])}
    all_codes = list(dict.fromkeys(list(soap_docs.keys()) + list(rest_docs.keys())))

    for code in all_codes:
        sd = soap_docs.get(code, {})
        rd = rest_docs.get(code, {})
        doc_name = sd.get('requiredDocumentName', '') or rd.get('requiredDocumentName', '') or code

        doc_result = {
            'code': code,
            'name': doc_name,
            'inSoap': code in soap_docs,
            'inRest': code in rest_docs,
            'meta': [], 'templateFields': [], 'indexes': []
        }

        if code in soap_docs and code in rest_docs:
            for mf in ['documentType', 'willBeScanned', 'printable', 'signatureRequired',
                       'minPage', 'maxPage', 'mandatory', 'sendEmail', 'fromCampaign',
                       'requiredDocumentCode', 'requiredDocumentName', 'canBeSentLater',
                       'barcodeCode', 'templateId']:
                sv, rv = sd.get(mf, ''), rd.get(mf, '')
                is_same = normalize(sv) == normalize(rv)
                doc_result['meta'].append({
                    'field': mf, 'soap': sv or '*(boş)*', 'rest': rv or '*(boş)*',
                    'status': 'same' if is_same else 'diff'
                })
                results['summary']['total'] += 1
                results['summary']['same' if is_same else 'diff'] += 1

            stf, rtf = sd.get('templateFields', {}), rd.get('templateFields', {})
            for tf in list(dict.fromkeys(list(stf.keys()) + list(rtf.keys()))):
                sv, rv = stf.get(tf, ''), rtf.get(tf, '')
                is_same = normalize(sv) == normalize(rv)
                doc_result['templateFields'].append({
                    'field': tf, 'soap': sv or '*(boş)*', 'rest': rv or '*(boş)*',
                    'status': 'same' if is_same else 'diff'
                })
                results['summary']['total'] += 1
                results['summary']['same' if is_same else 'diff'] += 1

            ss = sd.get('docStoreSystems', [{}])[0] if sd.get('docStoreSystems') else {}
            rs = rd.get('docStoreSystems', [{}])[0] if rd.get('docStoreSystems') else {}
            si, ri = ss.get('indexes', {}), rs.get('indexes', {})
            for idx in list(dict.fromkeys(list(si.keys()) + list(ri.keys()))):
                sv, rv = si.get(idx, ''), ri.get(idx, '')
                is_same = normalize(sv) == normalize(rv)
                doc_result['indexes'].append({
                    'field': idx, 'soap': sv or '*(boş)*', 'rest': rv or '*(boş)*',
                    'status': 'same' if is_same else 'diff'
                })
                results['summary']['total'] += 1
                results['summary']['same' if is_same else 'diff'] += 1

        results['documents'].append(doc_result)

    return results


# ========================== EXCEL EXPORT ==========================
def export_to_excel(comparison_data, rest_data):
    """Karşılaştırma sonuçlarını Excel'e aktarır (openpyxl ile)."""
    
    if not os.path.exists(EXPORTS_DIR):
        os.makedirs(EXPORTS_DIR)
        print(f"📁 Exports klasörü oluşturuldu: {EXPORTS_DIR}")
    
    # Sipariş numarasını al
    identifier = rest_data.get('identifier', 'unknown') if isinstance(rest_data, dict) else 'unknown'
    
    # Dosya adı: sipariş_numarası - tarih_saat.xlsx
    now = datetime.now()
    date_str = now.strftime("%Y%m%d_%H%M%S")
    filename = f"{identifier} - {date_str}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)
    
    print(f"📝 Excel oluşturuluyor: {filepath}")
    
    wb = openpyxl.Workbook()
    headers = ['Alan Adı', 'AS-IS (SOAP)', 'TO-BE (REST)', 'Durum']
    all_diffs = []
    
    # Sheet 1: Üst Seviye & Müşteri
    ws1 = wb.active
    ws1.title = 'Üst Seviye & Müşteri'
    set_col_widths(ws1)
    row = 1
    
    # Top Level
    add_title(ws1, row, '🔝 Üst Seviye (Top-Level) Alanlar')
    row += 1
    add_header_row(ws1, row, headers)
    row += 1
    for item in comparison_data.get('topLevel', []):
        is_diff = item['status'] == 'diff'
        status = '❌ Fark' if is_diff else '✅ Aynı'
        add_data_row(ws1, row, [item['field'], item['soap'], item['rest'], status], is_diff)
        if is_diff:
            all_diffs.append(('Üst Seviye', item['field'], item['soap'], item['rest']))
        row += 1
    row += 1
    
    # Customer
    add_title(ws1, row, '👤 Müşteri (Customer) Bilgileri')
    row += 1
    add_header_row(ws1, row, headers)
    row += 1
    for item in comparison_data.get('customer', []):
        is_diff = item['status'] == 'diff'
        status = '❌ Fark' if is_diff else '✅ Aynı'
        add_data_row(ws1, row, [item['field'], item['soap'], item['rest'], status], is_diff)
        if is_diff:
            all_diffs.append(('Müşteri', item['field'], item['soap'], item['rest']))
        row += 1
    
    # Document Sheets
    for i, doc in enumerate(comparison_data.get('documents', [])):
        if not doc.get('inSoap') or not doc.get('inRest'):
            continue
            
        doc_name = doc.get('name', doc.get('code', f'Doküman {i+1}'))
        sheet_name = doc_name[:31]  # Excel sheet name max 31 char
        
        # Aynı isimde sheet varsa numara ekle
        existing_names = [ws.title for ws in wb.worksheets]
        if sheet_name in existing_names:
            sheet_name = f"{sheet_name[:28]}_{i+1}"
        
        ws = wb.create_sheet(sheet_name)
        set_col_widths(ws)
        row = 1
        
        doc_type = ''
        template_id = ''
        for m in doc.get('meta', []):
            if m['field'] == 'documentType':
                doc_type = m['rest'] if m['rest'] != '*(boş)*' else m['soap']
            if m['field'] == 'templateId':
                template_id = m['rest'] if m['rest'] != '*(boş)*' else m['soap']
        
        add_title(ws, row, f"📄 Doküman {i+1}: {doc_name} ({doc_type} / {template_id})")
        row += 2
        
        # Meta
        add_subtitle(ws, row, 'Doküman Meta Bilgileri')
        row += 1
        add_header_row(ws, row, headers)
        row += 1
        for item in doc.get('meta', []):
            is_diff = item['status'] == 'diff'
            status = '❌ Fark' if is_diff else '✅ Aynı'
            add_data_row(ws, row, [item['field'], item['soap'], item['rest'], status], is_diff)
            if is_diff:
                all_diffs.append((doc_name, item['field'], item['soap'], item['rest']))
            row += 1
        
        # Indexes
        if doc.get('indexes'):
            row += 1
            add_subtitle(ws, row, 'DocStoreSystem - Document Indexes')
            row += 1
            add_header_row(ws, row, headers)
            row += 1
            for item in doc.get('indexes', []):
                is_diff = item['status'] == 'diff'
                status = '❌ Fark' if is_diff else '✅ Aynı'
                add_data_row(ws, row, [item['field'], item['soap'], item['rest'], status], is_diff)
                if is_diff:
                    all_diffs.append((doc_name, f"{item['field']} (index)", item['soap'], item['rest']))
                row += 1
        
        # Template Fields
        if doc.get('templateFields'):
            row += 1
            add_subtitle(ws, row, 'Document Template Fields')
            row += 1
            add_header_row(ws, row, headers)
            row += 1
            for item in doc.get('templateFields', []):
                is_diff = item['status'] == 'diff'
                status = '❌ Fark' if is_diff else '✅ Aynı'
                add_data_row(ws, row, [item['field'], item['soap'], item['rest'], status], is_diff)
                if is_diff:
                    all_diffs.append((doc_name, item['field'], item['soap'], item['rest']))
                row += 1
    
    # Summary Sheet - Tüm Farklar
    ws_summary = wb.create_sheet('ÖZET - Tüm Farklar')
    ws_summary.column_dimensions['A'].width = 8
    ws_summary.column_dimensions['B'].width = 30
    ws_summary.column_dimensions['C'].width = 40
    ws_summary.column_dimensions['D'].width = 45
    ws_summary.column_dimensions['E'].width = 45
    row = 1
    
    add_title(ws_summary, row, '🔴 ÖZET: Tüm Farklılıklar (Tüm Dokümanlardan)', 5)
    row += 2
    
    summary_headers = ['#', 'Doküman', 'Alan Adı', 'AS-IS (SOAP)', 'TO-BE (REST)']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = SUMMARY_HEADER_FILL
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
    row += 1
    
    for idx, (doc, field, asis, tobe) in enumerate(all_diffs, 1):
        for col, val in enumerate([idx, doc, field, asis, tobe], 1):
            cell = ws_summary.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
            cell.font = SUMMARY_VALUE_FONT if col >= 3 else NORMAL_FONT
        row += 1
    
    # Özet istatistikler
    row += 2
    summary = comparison_data.get('summary', {})
    ws_summary.cell(row=row, column=1, value="Sipariş No:").font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=identifier)
    row += 1
    ws_summary.cell(row=row, column=1, value="Rapor Tarihi:").font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=now.strftime("%d.%m.%Y %H:%M:%S"))
    row += 1
    ws_summary.cell(row=row, column=1, value="Toplam Alan:").font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=summary.get('total', 0))
    row += 1
    ws_summary.cell(row=row, column=1, value="Eşleşen:").font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=summary.get('same', 0))
    row += 1
    ws_summary.cell(row=row, column=1, value="Farklı:").font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=summary.get('diff', 0))
    
    wb.save(filepath)
    print(f"✅ Excel oluşturuldu: {filepath}")
    print(f"📊 Toplam fark sayısı: {len(all_diffs)}")
    
    return filepath, filename


# ========================== HTTP HANDLER ==========================
class CompareHandler(http.server.SimpleHTTPRequestHandler):
    def do_POST(self):
        content_length = int(self.headers['Content-Length'])
        post_data = self.rfile.read(content_length)
        
        if self.path == '/export-excel':
            self.handle_export_excel(post_data)
            return
        
        if self.path == '/compare':
            decoded_data = post_data.decode('utf-8')
            params = urllib.parse.parse_qs(decoded_data)

            soap_content = params.get('soap', [''])[0]
            rest_content = params.get('rest', [''])[0]

            if not soap_content.strip():
                result = {'error': 'SOAP (AS-IS) içeriği boş!'}
            elif not rest_content.strip():
                result = {'error': 'REST (TO-BE) içeriği boş!'}
            else:
                soap_data = parse_soap_xml(soap_content)
                rest_data = parse_rest_json(rest_content)
                result = compare_data(soap_data, rest_data)
                result['restData'] = rest_data

            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps(result, ensure_ascii=False).encode('utf-8'))
            return

    def handle_export_excel(self, post_data):
        try:
            data = json.loads(post_data.decode('utf-8'))
            comparison_data = data.get('comparison_data', {})
            rest_data = data.get('rest_data', {})
            
            filepath, filename = export_to_excel(comparison_data, rest_data)
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            response = json.dumps({'success': True, 'filename': filename, 'filepath': filepath})
            self.wfile.write(response.encode('utf-8'))
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.send_response(500)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode('utf-8'))

    def do_GET(self):
        if self.path.startswith('/download/'):
            filename = self.path[10:]
            filename = urllib.parse.unquote(filename)
            filepath = os.path.join(EXPORTS_DIR, filename)
            
            print(f"📥 Download isteği: {filename}")
            print(f"📂 Dosya yolu: {filepath}")
            
            if os.path.exists(filepath):
                file_size = os.path.getsize(filepath)
                print(f"✅ Dosya bulundu: {file_size} bytes")
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
                self.send_header('Content-Length', str(file_size))
                self.end_headers()
                
                with open(filepath, 'rb') as f:
                    self.wfile.write(f.read())
                return
            else:
                print(f"❌ Dosya bulunamadı: {filepath}")
                self.send_response(404)
                self.send_header('Content-Type', 'text/plain; charset=utf-8')
                self.end_headers()
                self.wfile.write(f"Dosya bulunamadı: {filename}".encode('utf-8'))
                return
        
        if self.path == '/' or self.path == '/index.html':
            self.send_response(200)
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.end_headers()
            html_path = os.path.join(SCRIPT_DIR, 'index.html')
            with open(html_path, 'r', encoding='utf-8') as f:
                self.wfile.write(f.read().encode('utf-8'))
        else:
            super().do_GET()


class ReuseAddrTCPServer(socketserver.TCPServer):
    allow_reuse_address = True


def main():
    print(f"📂 Script dizini: {SCRIPT_DIR}")
    print(f"📂 Exports dizini: {EXPORTS_DIR}")
    
    with ReuseAddrTCPServer(("", PORT), CompareHandler) as httpd:
        print(f"🚀 SOAP vs REST Karşılaştırma Aracı")
        print(f"📍 Tarayıcıda aç: http://localhost:{PORT}")
        print(f"❌ Durdurmak için: Ctrl+C")
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\n👋 Sunucu kapatıldı.")


if __name__ == '__main__':
    main()

# HTML template içinde (Python string olarak):

EXPORT_BUTTON_HTML = '''
<div id="export-section" style="display: none; margin-top: 20px; padding: 15px; background: #f5f5f5; border-radius: 8px;">
    <button id="exportBtn" onclick="exportToCSV()" 
            style="background-color: #217346; color: white; padding: 12px 24px; 
                   border: none; border-radius: 5px; cursor: pointer; font-size: 16px;
                   display: inline-flex; align-items: center; gap: 8px;">
        📊 Excel'e Aktar (CSV)
    </button>
    <span id="exportStatus" style="margin-left: 15px; font-size: 14px;"></span>
</div>

<script>
let lastComparisonData = null;
let lastTobeData = null;

// Karşılaştırma tamamlandığında çağrılacak
function onCompareComplete(comparisonData, tobeData) {
    lastComparisonData = comparisonData;
    lastTobeData = tobeData;
    document.getElementById('export-section').style.display = 'block';
}

function exportToCSV() {
    if (!lastComparisonData) {
        alert('Önce karşılaştırma yapmalısınız!');
        return;
    }
    
    const btn = document.getElementById('exportBtn');
    const status = document.getElementById('exportStatus');
    
    btn.disabled = true;
    btn.innerHTML = '⏳ Aktarılıyor...';
    status.textContent = '';
    
    fetch('/export-csv', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({
            comparison_data: lastComparisonData,
            tobe_data: lastTobeData
        })
    })
    .then(response => response.json())
    .then(data => {
        if (data.success) {
            status.innerHTML = '✅ Başarılı! <a href="/download/' + data.filename + '" download style="color: #217346; font-weight: bold;">📥 ' + data.filename + '</a>';
        } else {
            status.innerHTML = '❌ Hata: ' + data.error;
            status.style.color = 'red';
        }
    })
    .catch(error => {
        status.innerHTML = '❌ Hata: ' + error;
        status.style.color = 'red';
    })
    .finally(() => {
        btn.disabled = false;
        btn.innerHTML = '📊 Excel\\'e Aktar (CSV)';
    });
}
</script>
'''